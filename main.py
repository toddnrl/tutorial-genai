from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "자리배치"

# 열 너비 / 행 높이
for col in range(1, 19):
    ws.column_dimensions[get_column_letter(col)].width = 12
for row in range(1, 13):
    ws.row_dimensions[row].height = 26

# 스타일
beige_fill = PatternFill("solid", fgColor="F3E7C6")
blue_fill = PatternFill("solid", fgColor="B9C9E8")
center = Alignment(horizontal="center", vertical="center")

def set_cell(r, c, value="", fill=None, border_color="000000", bold=False, font_size=14, merge=None):
    if merge:
        ws.merge_cells(start_row=r, start_column=c, end_row=merge[0], end_column=merge[1])

    cell = ws.cell(r, c, value)
    cell.alignment = center
    cell.font = Font(name="Malgun Gothic", size=font_size, bold=bold)

    if fill:
        cell.fill = fill

    side = Side(style="thin", color=border_color)
    border = Border(left=side, right=side, top=side, bottom=side)

    if merge:
        for rr in range(r, merge[0] + 1):
            for cc in range(c, merge[1] + 1):
                ws.cell(rr, cc).border = border
                ws.cell(rr, cc).alignment = center
                ws.cell(rr, cc).font = Font(name="Malgun Gothic", size=font_size, bold=bold)
                if fill:
                    ws.cell(rr, cc).fill = fill
    else:
        cell.border = border

# 제목
set_cell(1, 8, "가족 피시방", bold=True, font_size=16, merge=(1, 10))

# 왼쪽 영역
set_cell(2, 1, "", border_color="2F64FF", merge=(2, 2))
set_cell(3, 1, "조윤", fill=beige_fill, border_color="FF2F92", merge=(3, 2))
set_cell(4, 1, "권대현", fill=beige_fill, merge=(4, 2))

for i, name in enumerate(["조은결", "김성호", "한혜경", "유지산"], start=3):
    set_cell(i, 3, name, fill=beige_fill, merge=(i, 4))

# 가운데 영역
for i, name in enumerate(["이춘우", "김민정", "진미경", "김민정"], start=3):
    set_cell(i, 7, name, fill=beige_fill, merge=(i, 8))

for i, name in enumerate(["최준호", "박지우", "박상우", "이종혁", "심경호"], start=3):
    set_cell(i, 9, name, fill=beige_fill, merge=(i, 10))

for i, name in enumerate(["이서우", "박은빈", "조용민", "김유선", "이태경"], start=3):
    set_cell(i, 11, name, fill=beige_fill, merge=(i, 12))

# 오른쪽 영역
for i, name in enumerate(["황동건", "원가희", "김선희", "박초롱", "채란"], start=3):
    set_cell(i, 15, name, fill=beige_fill, merge=(i, 16))

for i, name in enumerate(["이상욱", "신혜진", "한은태", "박준하", "이민지"], start=3):
    set_cell(i, 17, name, fill=beige_fill, merge=(i, 18))

set_cell(8, 17, "", border_color="FF2F92", merge=(8, 18))

# 아래쪽
set_cell(10, 1, "", border_color="D92B2B", merge=(10, 2))
set_cell(10, 5, "강사", fill=blue_fill, merge=(10, 6))
set_cell(10, 9, "앞쪽 스크린", bold=True, merge=(10, 11))

# 파일 저장
output_file = "seating_layout_like_image.xlsx"
wb.save(output_file)
print(f"저장 완료: {output_file}")